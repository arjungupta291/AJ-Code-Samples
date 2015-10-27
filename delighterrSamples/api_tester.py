import openpyxl
import random
from delighterr_functionality import Contact

def profile_percentage(contact):
	sections_filled = 0
	if contact.bio['name'] and (contact.bio['age'] or contact.bio['gender']):
		sections_filled += 1
	if contact.location_info['currently_lives']:
		sections_filled += 1
	if contact.location_info['previously_lived']:
		sections_filled += 1
	if contact.education_info['summary']:
		sections_filled += 1
	if contact.work_info['summary']:
		sections_filled += 1
	if contact.social_media['accounts']:
		sections_filled += 1
	return round(sections_filled / 6.0, 2) * 100

def has_parameters(row_int):
	first_name = 'A' + str(row_int)
	last_name = 'C' + str(row_int)
	email = 'O' + str(row_int)
	if ws[first_name].value and ws[last_name].value and ws[email].value:
		if len(ws[first_name].value) >= 2 and len(ws[last_name].value) >= 3: 
			return True
	return False

def pfilter(values):
	array = []
	for item in values:
		if has_parameters(item):
			array.append(item)
	return array

def random_selector(int1, int2):
	values_picked = [random.randint(int1, int2) for _ in range(200)]
	filtered_values = pfilter(values_picked)
	return filtered_values


def contact_selector(excel_wksht):
	random_values = random_selector(10,3000)
	contacts = []
	for num in random_values:
		first_name = ((excel_wksht['A' + str(num)]).value).encode("utf-8")
		last_name = ((excel_wksht['C' + str(num)]).value).encode("utf-8")
		email = ((excel_wksht['O' + str(num)]).value).encode("utf-8")
		person = Contact(first_name, last_name, email)
		contacts.append(person)
	return contacts

def fullcontact_tester(contacts):
	percentages = {}
	for contact in contacts:
		contact.auto_populate_fullcontact()
		has_linkedin = contact.has_linkedin()
		percent_complete = profile_percentage(contact)
		percentages[contact.emails[0].address] = [percent_complete, has_linkedin]
	return percentages

def pipl_tester(contacts):
	percentages = {}
	for contact in contacts:
		contact.auto_populate_pipl()
		has_linkedin = contact.has_linkedin()
		percent_complete = profile_percentage(contact)
		percentages[contact.emails[0].address] = [percent_complete, has_linkedin]
	return percentages

def compare(contacts):
	pipl = pipl_tester(contacts)
	fullcontact = fullcontact_tester(contacts)
	return pipl, fullcontact

def linkedin_percentage(dictp):
	profiles = 0.0
	linkedin = 0.0
	for person in dictp:
		profiles += 1
		if dictp[person][1][0]:
			linkedin += 1
	return round(linkedin / profiles, 2) * 100

def completion_average(dictp):
	profiles = 0.0
	complete = 0.0
	for person in dictp:
		profiles += 1
		complete += dictp[person][0]
	return round(complete / profiles, 2) 

def store(contacts):
	pipl, fullcontact = compare(contacts)
	file = open("contactStats.txt", "a")
	for email in pipl:
		file.write(email + " pipl: (" + "percent: " + str(pipl[email][0]) + " linkedin: " + 
				   str(pipl[email][1][0]) + ")" + " fullcontact: (" + "percent: " + str(fullcontact[email][0]) + " linkedin: " + 
				   str(fullcontact[email][1][0]) + ")")
		file.write( "\n")
	file.write("fullcontact profile completion percentage: " + str(completion_average(fullcontact)))
	file.write( "\n")
	file.write("pipl profile completion percentage: " + str(completion_average(pipl)))
	file.write( "\n")
	file.write("fullcontact linkedin percentage: " + str(linkedin_percentage(fullcontact)))
	file.write("\n")
	file.write("pipl linkedin percentage: " + str(linkedin_percentage(pipl)))
	file.close()


wb = openpyxl.load_workbook('contacts.xlsx')

ws = wb["contacts"]


if __name__ == '__main__':
	contacts = contact_selector(ws)
	print "storing"
	store(contacts)